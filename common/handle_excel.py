import openpyxl


class HandleExcel:
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def data_read(self):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        row = list(sheet.rows)
        title = [i.value for i in row[0]]
        cases = []
        for item in row[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases

    def data_write(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sheet = workbook[self.sheetname]
        sheet.cell(row=row, column=column, value=value)
        workbook.save(self.filename)

